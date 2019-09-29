import openpyxl as xl
from openpyxl.chart import BarChart,Reference
def sheet_process(filename):
    wb=xl.load_workbook(filename)
    sheet=wb['Sheet1']
    cell1=sheet['a1']# alternative is
    cell=sheet.cell(1,1)
    cell2=sheet['b1']
    # cell=sheet.cell(1,1)
    print(cell1.value+"\t"+cell2.value)
    print(sheet.max_row)#---Prints max no of rows in excel sheet
    cell3=sheet.cell(1,4)
    cell3.value='Corrected Price'
    print(sheet.cell(1,4).value)
    for row in range(2,sheet.max_row+1):
        cell=sheet.cell(row,3)
        print(cell.value)
        corrected_values=cell.value*0.9
        corrected_values_cells=sheet.cell(row,4)
        corrected_values_cells.value=corrected_values
    val=Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
    chart=BarChart()
    chart.add_data(val)
    sheet.add_chart(chart,'E2')
    wb.save(filename)
sheet_process('transactions.xlsx')